"""
Excel Report Generator
Creates a beautiful, color-coded Excel file with all leads + email drafts.
"""
import os
import logging
from datetime import datetime
from typing import List

import openpyxl
from openpyxl.styles import (
    Font, Fill, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR

logger = logging.getLogger(__name__)

# ── Colors ────────────────────────────────────────────────────────────────────
COLORS = {
    "header_bg":    "1a1a2e",   # dark navy
    "header_fg":    "FFFFFF",
    "none_bg":      "FF6B6B",   # red — no website
    "old_bg":       "FFD93D",   # yellow — old website
    "modern_bg":    "6BCB77",   # green — modern website
    "alt_row":      "F8F9FA",   # light grey alternate row
    "border":       "DEE2E6",
    "title_bg":     "667eea",   # purple (AstiScale brand)
}

COLUMNS = [
    ("Įmonė",               30),
    ("Vadovas",             22),
    ("Telefonas",           18),
    ("El. paštas",          30),
    ("Svetainė",            30),
    ("Miestas",             14),
    ("Industrija",          20),
    ("Svetainės statusas",  18),
    ("Siūlomos paslaugos",  35),
    ("Rekvizitai URL",      35),
    ("El. laiško juodraštis", 80),
    ("Skambučio scenarijus", 60),
    ("Pastabos",            25),
]


def save_excel(leads: List, date_str: str = None) -> str:
    """
    Save leads to Excel. Returns path to saved file.
    """
    if not date_str:
        date_str = datetime.now().strftime("%Y-%m-%d")

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filepath = os.path.join(OUTPUT_DIR, f"leads_{date_str}.xlsx")

    wb = openpyxl.Workbook()

    # ── Leads Sheet ───────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Potencialūs klientai"

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    title_cell = ws["A1"]
    title_cell.value = f"🎯 AstiScale — Potencialūs klientai  |  {date_str}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=COLORS["header_fg"])
    title_cell.fill = PatternFill("solid", fgColor=COLORS["title_bg"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Header row (row 2)
    header_fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    header_font = Font(name="Calibri", bold=True, size=10, color=COLORS["header_fg"])
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[2].height = 24
    ws.freeze_panes = "A3"

    # Data rows
    for row_idx, lead in enumerate(leads, start=3):
        _write_lead_row(ws, row_idx, lead)

    # Auto-filter
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}2"

    # ── Stats Sheet ───────────────────────────────────────────────────────────
    ws_stats = wb.create_sheet("Statistika")
    _write_stats_sheet(ws_stats, leads, date_str)

    # ── Email Drafts Sheet ────────────────────────────────────────────────────
    ws_emails = wb.create_sheet("El. laiškai (visi)")
    _write_emails_sheet(ws_emails, leads)

    wb.save(filepath)
    logger.info(f"Excel saved: {filepath}")
    return filepath


def _write_lead_row(ws, row_idx: int, lead):
    """Write one lead into a worksheet row."""
    ws_status = lead.website_status or "none"

    # Row color based on website status
    if ws_status == "none":
        row_fill = PatternFill("solid", fgColor=COLORS["none_bg"])
    elif ws_status == "old":
        row_fill = PatternFill("solid", fgColor=COLORS["old_bg"])
    else:
        row_fill = PatternFill("solid", fgColor=COLORS["modern_bg"])

    # Alternate light rows for modern ones only
    if ws_status == "modern" and row_idx % 2 == 0:
        row_fill = PatternFill("solid", fgColor=COLORS["alt_row"])

    from processors.service_recommender import build_service_summary, cold_call_script

    values = [
        lead.company_name,
        lead.vadovas,
        lead.phone,
        lead.email,
        lead.website,
        lead.city,
        lead.industry,
        _status_label(ws_status),
        build_service_summary(lead.recommended_services),
        lead.rekvizitai_url,
        lead.email_draft,
        cold_call_script(lead),
        lead.notes,
    ]

    base_font = Font(name="Calibri", size=9)
    link_font = Font(name="Calibri", size=9, color="0563C1", underline="single")

    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value or "")
        cell.fill = row_fill
        cell.font = base_font
        cell.alignment = Alignment(
            vertical="top", wrap_text=(col_idx in (11, 12)),  # wrap email + script columns
        )
        # Make URLs clickable
        if col_idx in (5, 10) and value and value.startswith("http"):
            cell.hyperlink = value
            cell.font = link_font

    ws.row_dimensions[row_idx].height = 60 if lead.email_draft else 20


def _status_label(status: str) -> str:
    return {
        "none": "❌ Nėra svetainės",
        "old": "⚠️ Sena svetainė",
        "modern": "✅ Moderni svetainė",
    }.get(status, status)


def _write_stats_sheet(ws, leads: List, date_str: str):
    """Summary statistics."""
    ws["A1"] = "📊 Statistika"
    ws["A1"].font = Font(bold=True, size=14)

    total = len(leads)
    no_site = sum(1 for l in leads if l.website_status == "none")
    old_site = sum(1 for l in leads if l.website_status == "old")
    modern = sum(1 for l in leads if l.website_status == "modern")
    has_email = sum(1 for l in leads if l.email)
    has_phone = sum(1 for l in leads if l.phone)
    has_vadovas = sum(1 for l in leads if l.vadovas)

    stats = [
        ("Data", date_str),
        ("Iš viso leadų", total),
        ("Nėra svetainės ❌", no_site),
        ("Sena svetainė ⚠️", old_site),
        ("Moderni svetainė ✅", modern),
        ("Turi el. paštą", has_email),
        ("Turi telefoną", has_phone),
        ("Žinomas vadovas", has_vadovas),
    ]

    for i, (label, value) in enumerate(stats, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20

    # Industry breakdown
    ws["A13"] = "Pagal industriją:"
    ws["A13"].font = Font(bold=True)

    from collections import Counter
    industry_counts = Counter(l.industry for l in leads)
    for i, (ind, cnt) in enumerate(industry_counts.most_common(), start=14):
        ws.cell(row=i, column=1, value=ind)
        ws.cell(row=i, column=2, value=cnt)


def _write_emails_sheet(ws, leads: List):
    """Full email drafts for easy copy-paste."""
    ws["A1"] = "📧 El. laiškų juodraščiai"
    ws["A1"].font = Font(bold=True, size=14)
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 100

    row = 3
    for lead in leads:
        if not lead.email_draft:
            continue

        ws.cell(row=row, column=1, value="Įmonė:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=lead.company_name)
        row += 1
        ws.cell(row=row, column=1, value="Gavėjas:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=lead.email or "— nežinomas —")
        row += 1
        ws.cell(row=row, column=1, value="Laiškas:").font = Font(bold=True)
        email_cell = ws.cell(row=row, column=2, value=lead.email_draft)
        email_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = max(120, lead.email_draft.count("\n") * 15)
        row += 2
        # Separator
        ws.cell(row=row, column=1, value="─" * 40)
        row += 2
